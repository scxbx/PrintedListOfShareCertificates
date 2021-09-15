import os

import openpyxl
from openpyxl.styles import Alignment, Side, Border
from openpyxl.styles.borders import BORDER_THIN


def convert_to_number(letter, column_a=1):
    """
    字母列号转数字

    :param column_a: 你希望A列是第几列(0 or 1)? 默认1
    :return: int
    """
    ab = '_ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    letter0 = letter.upper()
    w = 0
    for _ in letter0:
        w *= 26
        w += ab.find(_)
    return w - 1 + column_a


class MsgFiller:
    def __init__(self, committee_list, sample_path, out_path):
        self.__sample_path = sample_path
        self.__out_path = out_path
        self.__committee_list = committee_list

    def fill(self):
        if not os.path.exists(self.__sample_path) or not self.__sample_path.endswith('.xlsx'):
            print('模板：{} 错误'.format(self.__sample_path))
            return
        wb = openpyxl.load_workbook(self.__sample_path, data_only=True)
        ws = wb['表2.农村集体经济组织股权证打印清单']

        start_row = 3
        count = 0

        for index, family in enumerate(self.__committee_list):
            row_c = start_row + index
            ws.cell(row_c, 1).value = index + 1
            ws.cell(row_c, 2).value = 1
            ws.cell(row_c, 3).value = family.org_name
            ws.cell(row_c, 4).value = family.credit_code
            ws.cell(row_c, 5).value = family.certificate
            ws.cell(row_c, 6).value = family.master_name
            ws.cell(row_c, 7).value = family.member_num
            count += family.member_num

        row_sum = start_row + len(self.__committee_list)
        ws.cell(row_sum, 1).value = '合计'
        ws.cell(row_sum, 6).value = '合计'
        end_row = row_sum - 1
        formula = '=SUM(G{}:G{})'.format(start_row, end_row)
        formula2 = '=SUM(B{}:B{})'.format(start_row, end_row)
        ws.cell(row_sum, 7).value = formula
        ws.cell(row_sum, 2).value = formula2
        ws.merge_cells(start_row=row_sum, end_row=row_sum, start_column=3, end_column=5)

        # Style
        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        for i in range(start_row, row_sum+1):
            ws.row_dimensions[i].height = 25.1
            for j in range(1, 8):
                ws.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(i, j).border = thin_border

        wb.save(self.__out_path)
